import numpy as np
import matplotlib.cm as cm
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
import math
import openpyxl
from PIL import Image

# appromiate solution
# Load the workbook
workbook1 = openpyxl.load_workbook('plot_test.xlsx')
workbook2 = openpyxl.load_workbook('plot_train.xlsx')
# Select the worksheet
worksheet1 = workbook1['Sheet1']
worksheet2 = workbook2['Sheet1']
# Create empty lists to store the data
list1 = []
list2 = []
list3 = []
list4 = []
lista = []
listb = []
listc = []
listd = []
# Loop through the rows in the worksheet and append the data to the appropriate list
for row in worksheet1.iter_rows(values_only=True):
    list1.append(row[0])
    list2.append(row[1])
    list3.append(row[2])
    list4.append(row[3])

for row in worksheet2.iter_rows(values_only=True):
    lista.append(row[0])
    listb.append(row[1])
    listc.append(row[2])
    listd.append(row[3])

plt.figure()
plt.rcParams['xtick.direction'] = 'in'
plt.rcParams['ytick.direction'] = 'in'
plt.rcParams['xtick.top'] = True
plt.rcParams['ytick.right'] = True

plt.semilogy(list1, list2, marker='*', label='Train', color='blue')
#plt.fill_between(list1, list3, list4, alpha=0.5, color=(142/255, 207/255, 201/255))
plt.semilogy(lista, listb, marker='<', label='Test', color='orange')

#plt.fill_between(lista, listc, listd, alpha=0.5, color=(255/255, 190/255, 122/255))
plt.xlabel("Steps")
plt.ylabel("MSE")
plt.legend()

plt.savefig('MSE.png', dpi=800)
plt.show()